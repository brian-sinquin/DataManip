"""
DataTable study for numerical data manipulation.

Provides tabular data editing with formula columns, units,
and uncertainty tracking.
"""

from __future__ import annotations
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
import numpy as np
import sympy as sp
import concurrent.futures

from core.study import Study
from core.data_object import DataObject
from core.formula_engine import FormulaEngine
from core.undo_manager import UndoManager, UndoAction, ActionType, UndoContext
from utils.uncertainty_propagation import UncertaintyPropagator


class ColumnType:
    """Column type enumeration."""
    DATA = "data"
    CALCULATED = "calculated"
    DERIVATIVE = "derivative"  # Numerical differentiation
    RANGE = "range"  # Generated ranges
    UNCERTAINTY = "uncertainty"  # Propagated uncertainties
    INTERPOLATION = "interpolation"  # Future


class DataTableStudy(Study):
    """Study for numerical data tables.
    
    Features:
    - Multiple column types (data, calculated, derived, etc.)
    - Formula evaluation with dependency tracking
    - Unit tracking and conversion
    - Uncertainty propagation
    - Variables/constants
    
    Attributes:
        table: Main DataObject containing the table
        column_metadata: Metadata for each column (type, formula, unit)
        variables: Global variables/constants
        formula_engine: Formula evaluator
    """
    
    def __init__(self, name: str, workspace=None):
        """Initialize DataTable study.
        
        Args:
            name: Study name
            workspace: Reference to parent workspace for variables access
        """
        super().__init__(name)
        
        # Reference to workspace for variables/constants
        self.workspace = workspace
        
        # Create main table DataObject
        self.table = DataObject.empty(name="main_table")
        self.add_data_object(self.table)
        
        # Column metadata: {col_name: {type, formula, unit, uncertainty}}
        self.column_metadata: Dict[str, Dict[str, Any]] = {}
        
        # Formula engine
        self.formula_engine = FormulaEngine()
        
        # Undo/Redo manager
        self.undo_manager = UndoManager(max_history=50)
        
        # Dirty flag tracking for lazy evaluation
        self._dirty_columns: set = set()
        self._dependency_graph: Dict[str, set] = {}  # col -> dependents
        self._auto_recalc: bool = True  # Auto-recalc on data changes
    
    def get_type(self) -> str:
        """Get study type identifier."""
        return "data_table"
    
    # ========================================================================
    # Dirty Flag Tracking & Dependencies
    # ========================================================================
    
    def mark_dirty(self, column_name: str):
        """Mark column and its dependents as dirty."""
        if column_name not in self.table.data.columns:
            return
        
        # Mark this column
        self._dirty_columns.add(column_name)
        
        # Mark all dependents recursively
        to_mark = [column_name]
        while to_mark:
            col = to_mark.pop()
            if col in self._dependency_graph:
                for dependent in self._dependency_graph[col]:
                    if dependent not in self._dirty_columns:
                        self._dirty_columns.add(dependent)
                        to_mark.append(dependent)
    
    def mark_clean(self, column_name: str):
        """Mark column as clean after recalculation."""
        self._dirty_columns.discard(column_name)
    
    def is_dirty(self, column_name: str) -> bool:
        """Check if column needs recalculation."""
        return column_name in self._dirty_columns
    
    def get_dirty_columns(self) -> set:
        """Get all dirty columns."""
        return self._dirty_columns.copy()
    
    def _add_dependency(self, column: str, depends_on: str):
        """Register that column depends on depends_on."""
        if depends_on not in self._dependency_graph:
            self._dependency_graph[depends_on] = set()
        self._dependency_graph[depends_on].add(column)
    
    def _remove_dependency(self, column: str):
        """Remove all dependencies for column."""
        for deps in self._dependency_graph.values():
            deps.discard(column)
        self._dependency_graph.pop(column, None)
    
    def _update_dependencies(self, col_name: str, formula: Optional[str]):
        """Update dependency graph when column formula changes."""
        # Remove old dependencies
        self._remove_dependency(col_name)
        
        # Add new dependencies
        if formula:
            deps = self.formula_engine.extract_dependencies(formula)
            for dep in deps:
                if dep in self.table.data.columns:
                    self._add_dependency(col_name, dep)
    
    # ========================================================================
    # Column Management
    # ========================================================================
    
    def add_column(
        self,
        name: str,
        column_type: str = ColumnType.DATA,
        formula: Optional[str] = None,
        unit: Optional[str] = None,
        initial_data: Optional[pd.Series] = None,
        # Derivative-specific
        derivative_of: Optional[str] = None,
        with_respect_to: Optional[str] = None,
        order: int = 1,
        # Range-specific
        range_type: Optional[str] = None,  # "linspace", "arange", "logspace"
        range_start: Optional[float] = None,
        range_stop: Optional[float] = None,
        range_count: Optional[int] = None,
        range_step: Optional[float] = None,
        # Uncertainty-specific
        propagate_uncertainty: bool = False,
        uncertainty_reference: Optional[str] = None
    ):
        """Add column to table.
        
        Args:
            name: Column name
            column_type: Column type (DATA, CALCULATED, DERIVATIVE, RANGE, UNCERTAINTY)
            formula: Formula string (for calculated columns)
            unit: Physical unit
            initial_data: Initial column values
            derivative_of: Source column for derivative (DERIVATIVE type)
            with_respect_to: Independent variable column (DERIVATIVE type)
            order: Derivative order (1=first, 2=second, etc.)
            range_type: Range generation method (RANGE type)
            range_start: Start value (RANGE type)
            range_stop: Stop value (RANGE type)
            range_count: Number of points (linspace, logspace)
            range_step: Step size (arange)
            propagate_uncertainty: Auto-calculate uncertainty (CALCULATED type)
            uncertainty_reference: Parent column name (UNCERTAINTY type)
        """
        # Check if undo tracking is enabled BEFORE disabling
        should_track_undo = self.undo_manager.is_enabled()
        
        # Temporarily disable undo tracking to avoid recursive undo actions
        if should_track_undo:
            self.undo_manager.set_enabled(False)
        
        try:
            # Add column to DataObject
            if initial_data is not None:
                self.table.set_column(name, initial_data)
            else:
                self.table.add_column(name)
            
            # Store metadata
            self.column_metadata[name] = {
                "type": column_type,
                "formula": formula,
                "unit": unit,
                "uncertainty": None,
                "derivative_of": derivative_of,
                "with_respect_to": with_respect_to,
                "order": order,
                "range_type": range_type,
                "range_start": range_start,
                "range_stop": range_stop,
                "range_count": range_count,
                "range_step": range_step,
                "propagate_uncertainty": propagate_uncertainty,
                "uncertainty_reference": uncertainty_reference
            }
            
            # Handle by type
            if column_type == ColumnType.CALCULATED and formula:
                self.formula_engine.register_formula(name, formula)
                self._update_dependencies(name, formula)
                if len(self.table.data) > 0:
                    if self._auto_recalc:
                        self._recalculate_column(name)
                    else:
                        # Mark dirty if not auto-recalculating (batch mode)
                        self.mark_dirty(name)
                
                # Auto-create uncertainty column if requested
                if propagate_uncertainty:
                    uncert_name = f"{name}_u"
                    if uncert_name not in self.column_metadata:
                        uncert_values = self._calculate_propagated_uncertainty(name)
                        self.add_column(
                            uncert_name,
                            column_type=ColumnType.UNCERTAINTY,
                            unit=unit,  # Same unit as parent
                            initial_data=uncert_values,
                            uncertainty_reference=name
                        )
            
            elif column_type == ColumnType.DERIVATIVE:
                if len(self.table.data) > 0:
                    self._calculate_derivative(name)
            
            elif column_type == ColumnType.RANGE:
                self._generate_range(name)
        
        finally:
            # Re-enable undo tracking
            if should_track_undo:
                self.undo_manager.set_enabled(True)
                
                # Create undo action after column is fully set up
                # Capture final column data and metadata
                column_data = self.table.get_column(name).copy()
                metadata = self.column_metadata[name].copy()
                uncert_name = f"{name}_u" if propagate_uncertainty else None
                uncert_data = self.table.get_column(uncert_name).copy() if uncert_name and uncert_name in self.table.data.columns else None
                uncert_metadata = self.column_metadata.get(uncert_name, {}).copy() if uncert_name else None
                
                def undo_add():
                    with UndoContext(self.undo_manager, enabled=False):
                        self.remove_column(name)
                        # Also remove auto-created uncertainty column if it exists
                        if uncert_name and uncert_name in self.table.data.columns:
                            self.remove_column(uncert_name)
                
                def redo_add():
                    with UndoContext(self.undo_manager, enabled=False):
                        self.table.set_column(name, column_data.copy())
                        self.column_metadata[name] = metadata.copy()
                        # Restore formula registration if needed
                        if metadata.get("formula"):
                            self.formula_engine.register_formula(name, metadata["formula"])
                        # Restore uncertainty column if needed
                        if uncert_data is not None and uncert_metadata is not None:
                            self.table.set_column(uncert_name, uncert_data.copy())
                            self.column_metadata[uncert_name] = uncert_metadata.copy()
                
                action = UndoAction(
                    action_type=ActionType.ADD_COLUMN,
                    undo_func=undo_add,
                    redo_func=redo_add,
                    description=f"Add column '{name}'"
                )
                self.undo_manager.push(action)
    
    def add_columns_batch(self, columns: List[Dict[str, Any]]) -> List[str]:
        """Add multiple columns in one batch operation.
        
        Args:
            columns: List of dicts with keys: name, column_type, formula, 
                    initial_data, unit, uncertainty_formula, etc.
        
        Returns:
            List of added column names
        """
        added = []
        
        # Disable auto-recalculation during batch
        auto_recalc = self._auto_recalc
        self._auto_recalc = False
        
        try:
            for col_spec in columns:
                name = col_spec.get('name')
                if not name:
                    continue
                
                col_type = col_spec.get('column_type', ColumnType.DATA)
                formula = col_spec.get('formula')
                initial_data = col_spec.get('initial_data')
                unit = col_spec.get('unit', '')
                
                # Add column without auto-recalc
                self.add_column(
                    name, col_type, formula, initial_data, unit
                )
                added.append(name)
        
        finally:
            # Re-enable auto-recalc
            self._auto_recalc = auto_recalc
            
            # Recalculate all dirty columns once
            if auto_recalc:
                self._recalculate_dirty_columns()
        
        return added
    
    def remove_column(self, name: str):
        """Remove column from table.
        
        Args:
            name: Column name
        """
        # Save state for undo
        if self.undo_manager.is_enabled():
            column_data = self.table.get_column(name).copy()
            metadata = self.column_metadata.get(name, {}).copy()
            
            def undo_remove():
                self.table.add_column(name, column_data)
                if metadata:
                    self.column_metadata[name] = metadata
            
            def redo_remove():
                self.table.remove_column(name)
                if name in self.column_metadata:
                    del self.column_metadata[name]
            
            action = UndoAction(
                action_type=ActionType.REMOVE_COLUMN,
                undo_func=undo_remove,
                redo_func=redo_remove,
                description=f"Remove column '{name}'"
            )
            self.undo_manager.push(action)
        
        self.table.remove_column(name)
        if name in self.column_metadata:
            del self.column_metadata[name]
    
    def rename_column(self, old_name: str, new_name: str):
        """Rename a column.
        
        Args:
            old_name: Current column name
            new_name: New column name
        """
        # Save state for undo
        if self.undo_manager.is_enabled():
            def undo_rename():
                self.rename_column_internal(new_name, old_name)
            
            def redo_rename():
                self.rename_column_internal(old_name, new_name)
            
            action = UndoAction(
                action_type=ActionType.RENAME_COLUMN,
                undo_func=undo_rename,
                redo_func=redo_rename,
                description=f"Rename column '{old_name}' to '{new_name}'"
            )
            self.undo_manager.push(action)
        
        self.rename_column_internal(old_name, new_name)
    
    def rename_column_internal(self, old_name: str, new_name: str):
        """Internal rename without undo tracking.
        
        Args:
            old_name: Current column name
            new_name: New column name
        """
        # Rename in DataFrame
        self.table.data.rename(columns={old_name: new_name}, inplace=True)
        
        # Update metadata
        if old_name in self.column_metadata:
            self.column_metadata[new_name] = self.column_metadata.pop(old_name)
        
        # Update formulas that reference this column
        for col_name, metadata in list(self.column_metadata.items()):
            if col_name == new_name:
                # Skip the renamed column itself
                continue
            
            formula = metadata.get("formula")
            if formula and f"{{{old_name}}}" in formula:
                # Replace {old_name} with {new_name} in formulas
                new_formula = formula.replace(f"{{{old_name}}}", f"{{{new_name}}}")
                metadata["formula"] = new_formula
                # Re-register with formula engine if it's a calculated column
                if metadata.get("type") in (ColumnType.CALCULATED, ColumnType.DATA):
                    self.formula_engine.register_formula(col_name, new_formula)
        
        # Re-register renamed column's formula with formula engine (if it has one)
        renamed_metadata = self.column_metadata.get(new_name, {})
        formula = renamed_metadata.get("formula")
        if formula:
            self.formula_engine.register_formula(new_name, formula)
        
        # Recalculate dependent columns
        self.recalculate_all()
    
    def get_column_type(self, name: str) -> str:
        """Get column type.
        
        Args:
            name: Column name
            
        Returns:
            Column type string
        """
        return self.column_metadata.get(name, {}).get("type", ColumnType.DATA)
    
    def get_column_formula(self, name: str) -> Optional[str]:
        """Get column formula.
        
        Args:
            name: Column name
            
        Returns:
            Formula string or None
        """
        return self.column_metadata.get(name, {}).get("formula")
    
    def get_column_unit(self, name: str) -> Optional[str]:
        """Get column unit.
        
        Args:
            name: Column name
            
        Returns:
            Unit string or None
        """
        return self.column_metadata.get(name, {}).get("unit")
    
    # ========================================================================
    # Formula Evaluation
    # ========================================================================
    
    def _generate_range(self, name: str):
        """Generate range data for a column.
        
        Args:
            name: Range column name
        """
        meta = self.column_metadata[name]
        range_type = meta.get("range_type")
        start = meta.get("range_start")
        stop = meta.get("range_stop")
        count = meta.get("range_count")
        step = meta.get("range_step")
        
        if not range_type:
            return
        
        # Generate data based on type
        data = None
        if range_type == "linspace":
            if start is not None and stop is not None and count is not None:
                data = np.linspace(start, stop, count)
        elif range_type == "arange":
            if start is not None and stop is not None:
                if step is None:
                    step = 1.0
                data = np.arange(start, stop, step)
        elif range_type == "logspace":
            if start is not None and stop is not None and count is not None:
                data = np.logspace(start, stop, count)
        
        if data is None:
            return
        
        # Store result - handle different lengths
        if len(self.table.data) == 0:
            # First column - initialize DataFrame
            self.table.data = pd.DataFrame({name: data})
        else:
            # Extend or truncate table to match new column length
            current_len = len(self.table.data)
            new_len = len(data)
            
            if new_len > current_len:
                # Extend table
                self.table.data = pd.concat([
                    self.table.data,
                    pd.DataFrame(index=range(current_len, new_len))
                ], ignore_index=True)
            
            # Now add the column (will fill NaN if shorter than table)
            if new_len < len(self.table.data):
                # Pad with NaN
                padded_data = np.full(len(self.table.data), np.nan)
                padded_data[:new_len] = data
                self.table.data[name] = padded_data
            else:
                self.table.data[name] = data
    
    def _calculate_derivative(self, name: str):
        """Calculate numerical derivative for a column.
        
        Args:
            name: Derivative column name
        """
        meta = self.column_metadata[name]
        y_col = meta.get("derivative_of")
        x_col = meta.get("with_respect_to")
        order = meta.get("order", 1)
        
        if not y_col or not x_col:
            return
        
        # Get data
        y = self.table.get_column(y_col).values
        x = self.table.get_column(x_col).values
        
        # Calculate derivative using numpy gradient
        deriv = np.gradient(y, x)
        
        # Higher order derivatives
        for _ in range(1, order):
            deriv = np.gradient(deriv, x)
        
        # Store result
        self.table.set_column(name, pd.Series(deriv))
    
    def _calculate_propagated_uncertainty(self, name: str) -> pd.Series:
        """Calculate propagated uncertainty for a calculated column.
        
        Uses symbolic differentiation to compute: δf = √(Σ(∂f/∂xᵢ · δxᵢ)²)
        
        Args:
            name: Calculated column name
            
        Returns:
            Series of propagated uncertainties
        """
        meta = self.column_metadata.get(name)
        if not meta or meta.get("type") != ColumnType.CALCULATED:
            return pd.Series([np.nan] * len(self.table.data))
        
        formula = meta.get("formula")
        if not formula:
            return pd.Series([np.nan] * len(self.table.data))
        
        # Extract dependencies from formula
        dependencies = self.formula_engine.extract_dependencies(formula)
        
        # Collect values and uncertainties for each dependency
        values = {}
        uncertainties = {}
        
        for dep in dependencies:
            if dep not in self.table.columns:
                continue
            
            values[dep] = self.table.get_column(dep)
            
            # Check if uncertainty column exists, use 0 if not
            uncert_col_name = f"{dep}_u"
            if uncert_col_name in self.table.columns:
                uncertainties[dep] = self.table.get_column(uncert_col_name)
            else:
                uncertainties[dep] = pd.Series([0.0] * len(self.table.data))
        
        # Use extracted uncertainty propagator
        workspace_constants = self.workspace.constants if self.workspace else {}
        
        return UncertaintyPropagator.calculate_propagated_uncertainty(
            formula=formula,
            dependencies=dependencies,
            values=values,
            uncertainties=uncertainties,
            workspace_constants=workspace_constants,
            math_functions=self.formula_engine._math_functions
        )
    
    def _recalculate_column(self, name: str, context: Optional[Dict[str, Any]] = None):
        """Recalculate a formula column.
        
        Args:
            name: Column name
            context: Optional pre-built context (for lazy recalculation)
        """
        # Skip if not dirty and auto-recalc disabled
        if not self._auto_recalc and not self.is_dirty(name):
            return
        
        formula = self.get_column_formula(name)
        if not formula:
            return
        
        # Build or use provided context
        if context is None:
            context = {}
        
            # Add all data columns (convert Series to numpy arrays for cleaner evaluation)
            for col_name in self.table.columns:
                col_data = self.table.get_column(col_name)
                # Convert to numpy array, replacing None with nan
                if isinstance(col_data, pd.Series):
                    arr = col_data.values
                    # Convert None to nan for proper numeric operations
                    arr = np.where(pd.isna(arr), np.nan, arr).astype(float)
                    context[col_name] = arr
                else:
                    context[col_name] = col_data
            
            # Build complete context including workspace constants
            workspace_constants = self.workspace.constants if self.workspace else None
            workspace_id = id(self.workspace) if self.workspace else None
            workspace_version = self.workspace._version if self.workspace else 0
            
            context = self.formula_engine.build_context_with_workspace(
                context, 
                workspace_constants,
                workspace_id=workspace_id,
                workspace_version=workspace_version
            )
        
        # Evaluate formula
        try:
            result = self.formula_engine.evaluate(formula, context)
        except Exception:
            # If evaluation fails, fill with NaN (silent - UI shows NaN)
            result = np.full(len(self.table.data), np.nan)
        
        # Store result (convert back to Series if needed)
        if not isinstance(result, pd.Series):
            result = pd.Series(result)
        
        self.table.set_column(name, result)
        
        # Mark as clean after successful calculation
        self.mark_clean(name)
        
        # Mark as clean after successful calculation
        self.mark_clean(name)
    
    def _recalculate_dirty_columns(self):
        """Recalculate only dirty columns, parallelizing independent ones."""
        dirty = self.get_dirty_columns()
        if not dirty:
            return
        
        # Build dependency levels for parallel execution
        levels = self._get_dependency_levels(dirty)
        
        # Build context once (reused for all columns)
        # Skip dirty columns since they'll be calculated
        context = {}
        for col_name in self.table.columns:
            if col_name in dirty:
                continue  # Skip dirty columns
            
            col_data = self.table.get_column(col_name)
            if isinstance(col_data, pd.Series):
                arr = col_data.values
                arr = np.where(pd.isna(arr), np.nan, arr).astype(float)
                context[col_name] = arr
            else:
                context[col_name] = col_data
        
        workspace_constants = self.workspace.constants if self.workspace else None
        workspace_id = id(self.workspace) if self.workspace else None
        workspace_version = self.workspace._version if self.workspace else 0
        
        context = self.formula_engine.build_context_with_workspace(
            context, workspace_constants, workspace_id, workspace_version
        )
        
        # Process each level (can parallelize within level)
        for level_cols in levels:
            if len(level_cols) == 1:
                # Single column - evaluate with shared context
                self._recalculate_column(level_cols[0], context)
                # Update context with new column data
                col_data = self.table.get_column(level_cols[0])
                if isinstance(col_data, pd.Series):
                    context[level_cols[0]] = col_data.values
            else:
                # Multiple independent columns - evaluate in parallel
                self._evaluate_columns_parallel(level_cols, context)
    
    def _evaluate_columns_parallel(self, columns: List[str], base_context: Dict[str, Any]):
        """Evaluate multiple independent columns in parallel."""
        formulas_to_eval = []
        col_names = []
        
        for col in columns:
            if col not in self.column_metadata:
                continue
            metadata = self.column_metadata[col]
            if metadata.get("type") == ColumnType.CALCULATED and metadata.get("formula"):
                formulas_to_eval.append((metadata["formula"], base_context.copy()))
                col_names.append(col)
        
        if not formulas_to_eval:
            return
        
        # Evaluate in parallel (limit to 4 workers for CPU-bound tasks)
        max_workers = min(4, len(formulas_to_eval))
        
        if max_workers > 1 and len(formulas_to_eval) > 1:
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [
                    executor.submit(self.formula_engine.evaluate, formula, ctx)
                    for formula, ctx in formulas_to_eval
                ]
                results = [f.result() for f in futures]
        else:
            # Single column or small batch - evaluate sequentially
            results = [self.formula_engine.evaluate(f, c) for f, c in formulas_to_eval]
        
        # Store results
        for col_name, result in zip(col_names, results):
            if result is not None:
                if not isinstance(result, pd.Series):
                    result = pd.Series(result)
                self.table.set_column(col_name, result)
                self.mark_clean(col_name)
                # Update base context for next level
                base_context[col_name] = result.values
    
    def _get_dependency_levels(self, columns: set) -> List[List[str]]:
        """Group columns into dependency levels for parallel execution."""
        levels = []
        remaining = columns.copy()
        
        while remaining:
            # Find columns with no dependencies in remaining set
            current_level = []
            for col in list(remaining):
                if col not in self.column_metadata:
                    remaining.remove(col)
                    continue
                
                metadata = self.column_metadata[col]
                formula = metadata.get("formula")
                
                if not formula:
                    current_level.append(col)
                    continue
                
                deps = self.formula_engine.extract_dependencies(formula)
                deps_in_remaining = [d for d in deps if d in remaining and d != col]
                
                if not deps_in_remaining:
                    current_level.append(col)
            
            if not current_level:
                # Circular dependency - break it
                current_level = [remaining.pop()]
            else:
                for col in current_level:
                    remaining.discard(col)
            
            levels.append(current_level)
        
        return levels
    
    def recalculate_all(self):
        """Recalculate all formula and derivative columns in dependency order."""
        # Recalculate calculated columns
        calc_columns = [
            name for name, meta in self.column_metadata.items()
            if meta.get("type") == ColumnType.CALCULATED
        ]
        for col_name in calc_columns:
            self._recalculate_column(col_name)
            
            # Update uncertainty column if it exists
            if self.column_metadata[col_name].get("propagate_uncertainty"):
                uncert_name = f"{col_name}_u"
                if uncert_name in self.column_metadata:
                    uncert_values = self._calculate_propagated_uncertainty(col_name)
                    self.table.set_column(uncert_name, uncert_values)
        
        # Recalculate derivative columns
        deriv_columns = [
            name for name, meta in self.column_metadata.items()
            if meta.get("type") == ColumnType.DERIVATIVE
        ]
        for col_name in deriv_columns:
            self._calculate_derivative(col_name)
    
    def on_data_changed(self, column_name: str):
        """Handle data change in a column.
        
        Args:
            column_name: Changed column name
        """
        # Get affected columns
        affected = self.formula_engine.get_calculation_order([column_name])
        
        # Recalculate affected columns
        for col_name in affected:
            self._recalculate_column(col_name)
            
            # Update uncertainty column if it exists
            if self.column_metadata.get(col_name, {}).get("propagate_uncertainty"):
                uncert_name = f"{col_name}_u"
                if uncert_name in self.column_metadata:
                    uncert_values = self._calculate_propagated_uncertainty(col_name)
                    self.table.set_column(uncert_name, uncert_values)
    
    # Variables are managed at workspace level via workspace.constants
    # Access via self.workspace.constants (type="constant")
    
    # ========================================================================
    # Row Operations
    # ========================================================================
    
    def add_rows(self, count: int):
        """Add rows to table.
        
        Args:
            count: Number of rows to add
        """
        current_rows = len(self.table.data)
        new_rows = pd.DataFrame(
            index=range(current_rows, current_rows + count),
            columns=self.table.columns
        )
        self.table.data = pd.concat([self.table.data, new_rows], ignore_index=True)
        
        # Recalculate formula columns
        self.recalculate_all()
    
    def remove_rows(self, indices: List[int]):
        """Remove rows from table.
        
        Args:
            indices: List of row indices to remove
        """
        self.table.data = self.table.data.drop(index=indices).reset_index(drop=True)
        
        # Recalculate formula columns
        self.recalculate_all()
    
    # ========================================================================
    # Serialization
    # ========================================================================
    
    def to_dict(self) -> Dict[str, Any]:
        """Export to dictionary."""
        base_dict = super().to_dict()
        base_dict.update({
            "column_metadata": self.column_metadata,
            # Variables are stored at workspace level, not study level
        })
        return base_dict
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any], workspace=None) -> DataTableStudy:
        """Create from dictionary."""
        study = cls(name=data["name"], workspace=workspace)
        
        # Restore table data
        if "data_objects" in data and "main_table" in data["data_objects"]:
            table_data = data["data_objects"]["main_table"]
            study.table = DataObject.from_dict(
                name=table_data["name"],
                data_dict=table_data["data"],
                **table_data.get("metadata", {})
            )
        
        # Restore metadata
        study.column_metadata = data.get("column_metadata", {})
        
        # Variables are managed at workspace level (not restored here)
        
        # Rebuild formula engine
        for col_name, meta in study.column_metadata.items():
            if meta.get("type") == ColumnType.CALCULATED and meta.get("formula"):
                study.formula_engine.register_formula(col_name, meta["formula"])
        
        return study
    
    def export_to_csv(self, filepath: str, include_metadata: bool = True) -> None:
        """Export table to CSV format.
        
        Args:
            filepath: Path to save CSV file
            include_metadata: If True, includes metadata as comment header
            
        Example:
            >>> study.export_to_csv("data.csv")
            >>> study.export_to_csv("data.csv", include_metadata=False)  # Data only
        """
        import csv
        
        df = self.table.data
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if include_metadata:
                # Write metadata header as comments
                f.write("# DataManip CSV Export\n")
                f.write(f"# Study: {self.name}\n")
                f.write(f"# Version: 0.2.0\n")
                f.write("#\n")
                f.write("# Column Metadata:\n")
                for col_name in self.table.columns:
                    meta = self.column_metadata.get(col_name, {})
                    col_type = meta.get("type", ColumnType.DATA)
                    f.write(f"#   {col_name}: type={col_type}")
                    
                    if meta.get("unit"):
                        f.write(f", unit={meta['unit']}")
                    if meta.get("formula"):
                        # Escape formula for CSV
                        formula = meta['formula'].replace(',', ';')
                        f.write(f", formula={formula}")
                    if col_type == ColumnType.DERIVATIVE:
                        f.write(f", derivative_of={meta.get('derivative_of', '')}")
                        f.write(f", with_respect_to={meta.get('with_respect_to', '')}")
                    if col_type == ColumnType.RANGE:
                        f.write(f", range_type={meta.get('range_type', '')}")
                    f.write("\n")
                f.write("#\n")
            
            # Write data using pandas
            df.to_csv(f, index=False)
    
    def import_from_csv(
        self, 
        filepath: str, 
        has_metadata: bool = True,
        delimiter: str = ",",
        encoding: str = "utf-8",
        decimal: str = ".",
        skip_rows: int = 0,
        header: int = 0
    ) -> None:
        """Import table from CSV format with advanced configuration.
        
        Args:
            filepath: Path to CSV file
            has_metadata: Whether file contains metadata comments
            delimiter: Column delimiter (default: ",")
            encoding: File encoding (default: "utf-8")
            decimal: Decimal separator for numbers (default: ".")
            skip_rows: Number of rows to skip at beginning
            header: Row number to use as column names (None for no header)
            
        Note:
            This replaces current table data. Complex metadata (formulas, derivatives)
            may need manual recreation after import.
            
        Example:
            >>> study.import_from_csv("data.csv")
            >>> study.import_from_csv("euro.csv", delimiter=";", decimal=",")
            >>> study.import_from_csv("simple.csv", has_metadata=False, header=None)
        """
        # Read CSV with advanced parameters
        df = pd.read_csv(
            filepath, 
            delimiter=delimiter,
            encoding=encoding,
            decimal=decimal,
            skiprows=skip_rows if skip_rows > 0 else None,
            header=header,
            comment='#' if has_metadata else None
        )
        
        # If no header, generate column names
        if header is None:
            df.columns = [f"Column_{i}" for i in range(len(df.columns))]
        
        # Clear existing table
        self.table = DataObject.from_dataframe("table", df)
        self.column_metadata.clear()
        self.formula_engine = FormulaEngine()
        
        # Create basic metadata for all columns
        for col_name in df.columns:
            self.column_metadata[col_name] = {
                "type": ColumnType.DATA,
                "unit": None
            }
        
        # TODO: Parse metadata from comments if has_metadata=True
        # For now, all columns imported as DATA type
    
    def export_to_excel(self, filepath: str, sheet_name: str = "Data") -> None:
        """Export table to Excel format.
        
        Args:
            filepath: Path to save Excel file (.xlsx)
            sheet_name: Name of the sheet to create
            
        Requires:
            openpyxl package (pip install openpyxl)
            
        Example:
            >>> study.export_to_excel("data.xlsx")
            >>> study.export_to_excel("data.xlsx", sheet_name="Experiment1")
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
        
        df = self.table.data
        
        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write metadata section
        ws['A1'] = 'DataManip Export'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Study: {self.name}'
        ws['A3'] = 'Version: 0.2.0'
        
        # Write column metadata
        row = 5
        ws[f'A{row}'] = 'Column Metadata:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Column'
        ws[f'B{row}'] = 'Type'
        ws[f'C{row}'] = 'Unit'
        ws[f'D{row}'] = 'Formula/Info'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', fill_type='solid')
        row += 1
        
        for col_name in self.table.columns:
            meta = self.column_metadata.get(col_name, {})
            ws[f'A{row}'] = col_name
            ws[f'B{row}'] = meta.get('type', ColumnType.DATA)
            ws[f'C{row}'] = meta.get('unit', '')
            
            # Formula or other info
            if meta.get('formula'):
                ws[f'D{row}'] = meta['formula']
            elif meta.get('type') == ColumnType.DERIVATIVE:
                ws[f'D{row}'] = f"d({meta.get('derivative_of', '')})/d({meta.get('with_respect_to', '')})"
            elif meta.get('type') == ColumnType.RANGE:
                ws[f'D{row}'] = meta.get('range_type', '')
            row += 1
        
        # Add spacing
        row += 2
        
        # Write data section
        ws[f'A{row}'] = 'Data:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Write column headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row, col_idx, col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='EEEEEE', fill_type='solid')
        row += 1
        
        # Write data rows
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row, col_idx, value)
            row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (AttributeError, TypeError):
                    # Skip cells that can't be converted to string
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + EXCEL_COLUMN_WIDTH_PADDING, EXCEL_MAX_COLUMN_WIDTH)
        
        # Save workbook
        wb.save(filepath)
    
    def import_from_excel(self, filepath: str, sheet_name: str = "Data") -> None:
        """Import table from Excel format.
        
        Args:
            filepath: Path to Excel file (.xlsx)
            sheet_name: Name of the sheet to read
            
        Note:
            Imports the data section only. Metadata needs manual recreation.
            
        Example:
            >>> study.import_from_excel("data.xlsx")
            >>> study.import_from_excel("data.xlsx", sheet_name="Experiment1")
        """
        # Use pandas to read Excel
        df = pd.read_excel(filepath, sheet_name=sheet_name, skiprows=None)
        
        # Find where data section starts (look for "Data:" marker)
        # For now, simple approach - assume data starts after empty rows
        # TODO: Smarter parsing to find data section
        
        # Clear existing table
        self.table = DataObject.from_dataframe("table", df)
        self.column_metadata.clear()
        self.formula_engine = FormulaEngine()
        
        # Create basic metadata
        for col_name in df.columns:
            self.column_metadata[col_name] = {
                "type": ColumnType.DATA,
                "unit": None
            }

